#!/usr/bin/env python3
"""Build the sparse residual-event calendar workbook from canonical artifacts."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REPORTS = Path("data_perp/reports")
DEFAULT_OUTPUT = REPORTS / "residual_sparse_event_calendar_overlay_recovery_20260714_v1.xlsx"

BLUE = "142E52"
LIGHT_BLUE = "EAF1F8"
WHITE = "FFFFFF"
GRID = "D7DEE8"


def compact_composites(value: Any, limit: int = 240) -> str:
    if pd.isna(value):
        return ""
    names: list[str] = []
    for name in str(value).split("|"):
        if name and name not in names:
            names.append(name)
    result = " | ".join(names[:4])
    return result if len(result) <= limit else f"{result[:limit - 3]}..."


def iso_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, utc=True).dt.strftime("%Y-%m-%d")


def optional_number(value: Any) -> float | None:
    return None if pd.isna(value) else float(value)


def calendar_with_blocks() -> pd.DataFrame:
    calendar = pd.read_csv(
        REPORTS
        / "residual_episode_recognition_calendar_20260712_v1"
        / "calendar_recognized_vs_ignored.csv"
    )
    blocks = pd.read_csv(
        REPORTS
        / "residual_event_block_taxonomy_20260714_v7_full_mechanism_calendar"
        / "event_block_mechanism_calendar.csv"
    )
    calendar["_day"] = pd.to_datetime(calendar["day"], utc=True)
    blocks["_start"] = pd.to_datetime(blocks["event_start"], utc=True)
    blocks["_end"] = pd.to_datetime(blocks["event_end"], utc=True)

    out = calendar.copy()
    for name in ["event_block", "event_start", "event_end", "onset_primary_mechanism"]:
        out[name] = ""
    out["onset_mechanism_confident"] = "no"

    for (side, archetype), index in out.groupby(["side_name", "archetype_policy_key"]).groups.items():
        group = out.loc[index]
        local = blocks[
            (blocks["side_name"] == side)
            & (blocks["archetype_policy_key"] == archetype)
        ]
        for _, block in local.iterrows():
            mask = (group["_day"] >= block["_start"]) & (group["_day"] <= block["_end"])
            rows = group.index[mask]
            out.loc[rows, "event_block"] = block["event_block"]
            out.loc[rows, "event_start"] = block["_start"].strftime("%Y-%m-%d")
            out.loc[rows, "event_end"] = block["_end"].strftime("%Y-%m-%d")
            out.loc[rows, "onset_primary_mechanism"] = block["onset_primary_mechanism"]
            out.loc[rows, "onset_mechanism_confident"] = (
                "yes" if bool(block["onset_mechanism_confident"]) else "no"
            )

    out["date"] = out["_day"].dt.strftime("%Y-%m-%d")
    for column in ["recognized", "adverse_state_detected", "adverse_selected_state_detected", "material_extreme"]:
        out[column] = out[column].map({True: "yes", False: "no"}).fillna("no")
    out["legacy_composites"] = out["matching_composites"].map(compact_composites)
    return out[
        [
            "date", "event_start", "event_end", "event_block", "side_name", "archetype_policy_key",
            "rows", "mean_ev_after_1pct", "clean_exec_precision", "daily_neutral_z",
            "daily_ev_neutral_z", "persistence_strength", "large_event_strength",
            "adverse_event_rows", "adverse_tail_rows", "selected_adverse_tail_rows",
            "material_extreme", "recognized", "status", "recognition_sources", "legacy_composites",
            "best_composite_score", "onset_primary_mechanism", "onset_mechanism_confident",
            "evidence_scope", "uncaptured_reason",
        ]
    ]


def block_table() -> pd.DataFrame:
    audit = pd.read_csv(
        REPORTS / "residual_episode_detection_audit_20260714_v3" / "episode_recognition_audit.csv"
    )
    coverage = pd.read_csv(
        REPORTS
        / "residual_mechanism_lookalike_validation_20260714_v3_chronological"
        / "focus_event_chronological_coverage.csv"
    ).rename(
        columns={
            "status": "chronological_status",
            "valid_arms_available": "chrono_valid_arms",
            "arms_alerted": "chrono_alerts",
            "max_valid_risk": "chrono_max_valid_risk",
        }
    )
    keys = ["event_start", "event_end", "side_name", "archetype_policy_key", "event_block"]
    table = audit.merge(
        coverage[keys + ["chronological_status", "chrono_valid_arms", "chrono_alerts", "chrono_max_valid_risk"]],
        on=keys,
        how="left",
    )
    table["legacy_composites"] = table["matching_composites"].map(compact_composites)
    for column in ["event_start", "event_end"]:
        table[column] = iso_date(table[column])
    for column in ["onset_mechanism_confident", "cnn_top05_early_warning", "cnn_oos_eligible"]:
        table[column] = table[column].map({True: "yes", False: "no"}).fillna("no")
    table["chronological_status"] = table["chronological_status"].fillna("not_focus_episode")
    for column in ["chrono_valid_arms", "chrono_alerts"]:
        table[column] = table[column].fillna(0).astype(int)
    return table[
        [
            "event_start", "event_end", "event_days", "side_name", "archetype_policy_key", "event_block",
            "calendar_mean_ev", "calendar_mean_signed_surprise", "calendar_persistence_strength",
            "calendar_large_event_strength", "legacy_calendar_status", "legacy_recognition_rate",
            "legacy_composites", "max_composite_score", "onset_primary_mechanism",
            "onset_primary_mechanism_score", "onset_mechanism_margin", "onset_mechanism_confident",
            "cnn_top05_early_warning", "cnn_oos_eligible", "chronological_status", "chrono_valid_arms",
            "chrono_alerts", "chrono_max_valid_risk",
        ]
    ]


def focus_table(blocks: pd.DataFrame) -> pd.DataFrame:
    coverage = pd.read_csv(
        REPORTS
        / "residual_mechanism_lookalike_validation_20260714_v3_chronological"
        / "focus_event_chronological_coverage.csv"
    )
    for column in ["event_start", "event_end"]:
        coverage[column] = iso_date(coverage[column])
    focused = pd.read_csv(
        REPORTS
        / "unresolved_event_mechanism_discovery_20260714_v4_eventconsistency"
        / "focused_event_model_metrics.csv"
    )
    cnn = pd.read_csv(
        REPORTS / "residual_hard_period_cnn_20260714_v5_tcn" / "hard_period_cnn_oof_summary.csv"
    )
    keys = ["event_start", "event_end", "side_name", "archetype_policy_key", "event_block"]
    detail = blocks.merge(coverage, on=keys, how="inner", suffixes=("", "_chrono"))

    compact = focused[
        (focused["status"] == "ok")
        & (focused["top05_selected_rate"] > 0)
        & (focused["top05_selected_rate"] <= 0.10)
    ].sort_values(
        ["side_name", "archetype_policy_key", "held_event", "top05_lift", "top05_recall"],
        ascending=[True, True, True, False, False],
    )
    compact = compact.groupby(["side_name", "archetype_policy_key", "held_event"], as_index=False).first()
    compact = compact.rename(
        columns={
            "held_event": "event_block",
            "model": "research_best_model",
            "top05_lift": "research_top05_lift",
            "top05_fpr": "research_top05_fpr",
            "top05_recall": "research_top05_recall",
            "top05_precision": "research_top05_precision",
        }
    )
    detail = detail.merge(
        compact[
            [
                "side_name", "archetype_policy_key", "event_block", "research_best_model",
                "research_top05_lift", "research_top05_fpr", "research_top05_recall",
                "research_top05_precision",
            ]
        ],
        on=["side_name", "archetype_policy_key", "event_block"],
        how="left",
    )
    best_cnn = cnn.sort_values(
        ["side_name", "archetype_policy_key", "passes_top05_repetition_gate", "top05_mean_lift"],
        ascending=[True, True, False, False],
    ).groupby(["side_name", "archetype_policy_key"], as_index=False).first()
    best_cnn = best_cnn.rename(
        columns={
            "model": "sequence_best_model",
            "top05_mean_lift": "sequence_top05_lift",
            "top05_mean_fpr": "sequence_top05_fpr",
            "top05_mean_event_recall": "sequence_top05_recall",
            "passes_top05_repetition_gate": "sequence_passes_gate",
        }
    )
    detail = detail.merge(
        best_cnn[
            [
                "side_name", "archetype_policy_key", "sequence_best_model", "sequence_top05_lift",
                "sequence_top05_fpr", "sequence_top05_recall", "sequence_passes_gate",
            ]
        ],
        on=["side_name", "archetype_policy_key"],
        how="left",
    )
    detail["research_status"] = np.where(
        detail["research_best_model"].notna(),
        "nonchronological research evidence only",
        "no compact research detector",
    )
    detail["recovery_assessment"] = np.select(
        [
            detail["status"].eq("no_eligible_prior_chronological_detector"),
            detail["arms_alerted"].fillna(0).gt(0),
            detail["valid_arms_available"].fillna(0).gt(0),
        ],
        [
            "no prior lookalike support: no causal detector was eligible",
            "causal candidate alerted; requires repetition gate",
            "causal detector scored but did not alert",
        ],
        default="no validated causal recovery overlay",
    )
    for column in ["event_start", "event_end"]:
        detail[column] = iso_date(detail[column])
    for column in ["sequence_passes_gate"]:
        detail[column] = detail[column].map({True: "yes", False: "no"}).fillna("no")
    return detail[
        [
            "event_start", "event_end", "side_name", "archetype_policy_key", "event_block",
            "calendar_mean_ev", "calendar_mean_signed_surprise", "calendar_persistence_strength",
            "calendar_large_event_strength", "legacy_calendar_status", "legacy_recognition_rate",
            "onset_primary_mechanism", "onset_primary_mechanism_score", "onset_mechanism_margin",
            "sequence_best_model", "sequence_top05_lift", "sequence_top05_fpr", "sequence_top05_recall",
            "sequence_passes_gate", "research_status", "research_best_model", "research_top05_lift",
            "research_top05_fpr", "research_top05_recall", "research_top05_precision", "status",
            "arms_available", "valid_arms_available", "arms_alerted", "max_valid_risk",
            "alerting_families", "recovery_assessment",
        ]
    ]


def method_table(blocks: pd.DataFrame) -> pd.DataFrame:
    cnn = pd.read_csv(
        REPORTS / "residual_hard_period_cnn_20260714_v5_tcn" / "hard_period_cnn_oof_summary.csv"
    )
    chronological = pd.read_csv(
        REPORTS
        / "residual_mechanism_lookalike_validation_20260714_v3_chronological"
        / "mechanism_validation_summary.csv"
    )
    focused = pd.read_csv(
        REPORTS
        / "unresolved_event_mechanism_discovery_20260714_v4_eventconsistency"
        / "focused_event_model_metrics.csv"
    )
    rows: list[dict[str, Any]] = []
    status_counts = blocks["legacy_calendar_status"].value_counts().to_dict()
    rows.append(
        {
            "evidence_class": "legacy_calendar",
            "method": "legacy composite recognition",
            "side": "all",
            "archetype": "all",
            "mechanism": "multiple discovery composites",
            "evaluation_contract": "discovery-period descriptive calendar",
            "folds": None,
            "lift": None,
            "fpr": None,
            "recall": blocks["legacy_recognition_rate"].mean(),
            "precision": None,
            "support": len(blocks),
            "status": "research evidence",
            "activation": "inactive",
            "notes": (
                f"blocks={len(blocks)}; fully={status_counts.get('fully_recognized', 0)}; "
                f"partial={status_counts.get('partially_recognized', 0)}; "
                f"not={status_counts.get('not_recognized', 0)}"
            ),
        }
    )
    rows.append(
        {
            "evidence_class": "mechanism_taxonomy",
            "method": "onset mechanism classifier",
            "side": "all",
            "archetype": "all",
            "mechanism": "observable market mechanisms",
            "evaluation_contract": "descriptive taxonomy",
            "folds": None,
            "lift": None,
            "fpr": None,
            "recall": (blocks["onset_mechanism_confident"] == "yes").mean(),
            "precision": None,
            "support": len(blocks),
            "status": "diagnostic",
            "activation": "inactive",
            "notes": "Confident onset assignments require a sufficient score margin.",
        }
    )
    for _, row in cnn[cnn["passes_top05_repetition_gate"]].iterrows():
        rows.append(
            {
                "evidence_class": "sequence_detector",
                "method": row["model"],
                "side": row["side_name"],
                "archetype": row["archetype_policy_key"],
                "mechanism": "local hard-period early warning",
                "evaluation_contract": "chronological OOF top-5% budget",
                "folds": int(row["folds"]),
                "lift": row["top05_mean_lift"],
                "fpr": row["top05_mean_fpr"],
                "recall": row["top05_mean_event_recall"],
                "precision": row["top05_mean_precision"],
                "support": None,
                "status": "passes detector gate",
                "activation": "research only",
                "notes": f"hit_folds={int(row['top05_hit_folds'])}; not wired into policy",
            }
        )
    for _, row in chronological[chronological["valid_folds"] > 0].iterrows():
        rows.append(
            {
                "evidence_class": "lookalike_detector",
                "method": "regularized LGBM",
                "side": row["side_name"],
                "archetype": row["archetype_policy_key"],
                "mechanism": row["mechanism_family"],
                "evaluation_contract": "prior-lookalike chronological; frozen 95th percentile threshold",
                "folds": int(row["valid_folds"]),
                "lift": optional_number(row["mean_lift"]),
                "fpr": optional_number(row["mean_fpr"]),
                "recall": optional_number(row["mean_event_recall"]),
                "precision": None,
                "support": optional_number(row["mean_train_lookalike_blocks"]),
                "status": "passes repetition gate" if row["passes_repetition_gate"] else "repeatability failed",
                "activation": "inactive",
                "notes": (
                    f"positive_lift_folds={int(row['positive_lift_folds'])}; "
                    f"hit_recall_folds={int(row['hit_recall_folds'])}"
                ),
            }
        )
    grouped = focused.groupby(["model", "status"], dropna=False).agg(
        arms=("held_event", "size"),
        mean_lift=("top05_lift", "mean"),
        mean_fpr=("top05_fpr", "mean"),
        mean_recall=("top05_recall", "mean"),
        mean_precision=("top05_precision", "mean"),
    ).reset_index()
    for _, row in grouped.iterrows():
        rows.append(
            {
                "evidence_class": "focused_model_search",
                "method": row["model"],
                "side": "focus groups",
                "archetype": "focus episodes",
                "mechanism": "held-event research search",
                "evaluation_contract": "nonchronological leave-one-event-out; diagnostic only",
                "folds": None,
                "lift": optional_number(row["mean_lift"]),
                "fpr": optional_number(row["mean_fpr"]),
                "recall": optional_number(row["mean_recall"]),
                "precision": optional_number(row["mean_precision"]),
                "support": int(row["arms"]),
                "status": row["status"],
                "activation": "inactive",
                "notes": "Not valid for policy selection; used only to identify mechanism hypotheses.",
            }
        )
    return pd.DataFrame(rows)


def write_table(ws, frame: pd.DataFrame, name: str, widths: dict[str, int]) -> None:
    header_fill = PatternFill("solid", fgColor=BLUE)
    header_font = Font(color=WHITE, bold=True)
    thin = Side(style="thin", color=GRID)
    for column_index, column in enumerate(frame.columns, 1):
        cell = ws.cell(1, column_index, column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 2):
        for column_index, value in enumerate(values, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.0000"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(frame.shape[1])}{frame.shape[0] + 1}"
    table = Table(displayName=name, ref=ws.auto_filter.ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    for column_index, column in enumerate(frame.columns, 1):
        width = widths.get(column, min(max(len(column) + 2, 12), 24))
        ws.column_dimensions[get_column_letter(column_index)].width = width
    ws.row_dimensions[1].height = 34


def add_color_scale(ws, columns: list[str], frame: pd.DataFrame) -> None:
    if frame.empty:
        return
    for column in columns:
        if column not in frame.columns:
            continue
        index = list(frame.columns).index(column) + 1
        ref = f"{get_column_letter(index)}2:{get_column_letter(index)}{len(frame) + 1}"
        ws.conditional_formatting.add(
            ref,
            ColorScaleRule(start_type="min", start_color="F4CCCC", mid_type="percentile",
                           mid_value=50, mid_color="FFF2CC", end_type="max", end_color="D9EAD3"),
        )


def write_overview(ws, calendar: pd.DataFrame, blocks: pd.DataFrame, methods: pd.DataFrame) -> None:
    summary = [
        ("Sparse Event Calendar: Overlay and Recovery Evidence", "", ""),
        ("Purpose", "Sparse adverse-event calendar with performance, affected archetypes, and evidence for existing recovery/overlay methods.", ""),
        ("Important contract", "Legacy composites and focused discovery evidence are descriptive/research-only. Chronological prior-lookalike evidence is the causal standard. This workbook does not activate a policy overlay.", ""),
        ("Metric", "Value", "Interpretation"),
        ("Daily event cells", len(calendar), "Rows in the canonical sparse daily calendar."),
        ("Contiguous event blocks", len(blocks), "Side x archetype contiguous adverse periods."),
        ("Calendar period", f"{calendar['date'].min()} to {calendar['date'].max()}", "Current source coverage."),
        ("Fully legacy-recognized blocks", int((blocks["legacy_calendar_status"] == "fully_recognized").sum()), "Legacy discovery composites recognized every event day in the block."),
        ("Partially legacy-recognized blocks", int((blocks["legacy_calendar_status"] == "partially_recognized").sum()), "At least one, but not all, days recognized."),
        ("Unrecognized legacy blocks", int((blocks["legacy_calendar_status"] == "not_recognized").sum()), "No legacy composite recognition."),
        ("Confident onset mechanisms", int((blocks["onset_mechanism_confident"] == "yes").sum()), "Observable taxonomy assignment with sufficient margin."),
        ("Strict sequence detector passes", int((methods["status"] == "passes detector gate").sum()), "Chronological OOF top-5% repetition gate; research only."),
        ("Strict chronological lookalike passes", int((methods["status"] == "passes repetition gate").sum()), "No mechanism passed the repeatability gate."),
        ("Workbook tabs", "Event Blocks / Daily Calendar / Focus Episodes / Method Evidence / Contracts", "Use Event Blocks for period-level decisions; Daily Calendar for day-level performance."),
        ("Primary sources", "residual_episode_recognition_calendar_20260712_v1; residual_event_block_taxonomy_20260714_v7_full_mechanism_calendar; residual_episode_detection_audit_20260714_v3", "Full provenance recorded in Contracts."),
    ]
    fill = PatternFill("solid", fgColor=BLUE)
    header = Font(color=WHITE, bold=True)
    for row_index, row in enumerate(summary, 1):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row_index, col_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index in (1, 4):
                cell.fill = fill
                cell.font = header
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 76
    for row in range(1, len(summary) + 1):
        ws.row_dimensions[row].height = 31


def write_contracts(ws) -> None:
    rows = [
        ("Term", "Definition", "Evidence / constraint"),
        ("Sparse daily event cell", "One day x side x archetype record from the canonical adverse-event calendar.", "Performance fields include mean EV after 1% cost, clean precision, signed surprise, persistence, and adverse-tail counts."),
        ("Contiguous event block", "Adjacent daily cells within the same side x archetype grouped into one adverse period.", "Block performance is a summary of daily source fields."),
        ("Legacy recognition", "Existing composite/leaf/unsupervised calendar matched a daily event cell.", "Discovery-period descriptive evidence, not a policy activation criterion."),
        ("Onset mechanism", "Train-derived observable market-mechanism taxonomy assigned at event onset.", "Diagnostic taxonomy; confidence requires a sufficient score margin."),
        ("Sequence detector", "Causal CNN/TCN/LGBM detector scored with chronological OOF folds and a top-5% alert budget.", "A detector gate is not a live policy change. Results remain research-only until a frozen forward contract passes."),
        ("Focused model search", "RuleFit, Bayesian-rule-list, contrastive subgroup, recursive partition, LGBM, MLP, CNN, and TCN around held hard episodes.", "Nonchronological diagnostic evidence only; never used to select a production overlay."),
        ("Chronological lookalike validation", "Local side x archetype mechanism detector trained only on prior same-family blocks; frozen 95th-percentile train threshold.", "Causal evaluation standard for recovery overlays. No mechanism passed repetition gate in the current artifact."),
        ("Lift", "Adverse-event rate within alerted rows divided by baseline adverse-event rate.", "Higher is useful only alongside low FPR and stable chronological support."),
        ("FPR", "Fraction of non-adverse rows alerted.", "Low FPR alone is insufficient if recall/support are weak."),
        ("Activation status", "Whether evidence changes meta score, policy threshold, or live behavior.", "All rows in this workbook are research/diagnostic unless explicitly stated otherwise; no changes are activated by this report."),
        ("Daily source", "data_perp/reports/residual_episode_recognition_calendar_20260712_v1/calendar_recognized_vs_ignored.csv", "Canonical sparse daily event source."),
        ("Taxonomy source", "data_perp/reports/residual_event_block_taxonomy_20260714_v7_full_mechanism_calendar/event_block_mechanism_calendar.csv", "Contiguous blocks and onset mechanisms."),
        ("Audit source", "data_perp/reports/residual_episode_detection_audit_20260714_v3/episode_recognition_audit.csv", "Legacy recognition, taxonomy, and sequence availability."),
        ("Lookalike source", "data_perp/reports/residual_mechanism_lookalike_validation_20260714_v3_chronological/", "Chronological local detector evidence."),
    ]
    frame = pd.DataFrame(rows[1:], columns=rows[0])
    write_table(ws, frame, "ContractsTable", {"Term": 30, "Definition": 74, "Evidence / constraint": 82})


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    calendar = calendar_with_blocks()
    blocks = block_table()
    focus = focus_table(blocks)
    methods = method_table(blocks)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    write_overview(overview, calendar, blocks, methods)

    event_blocks = workbook.create_sheet("Event Blocks")
    write_table(event_blocks, blocks, "EventBlocksTable", {"archetype_policy_key": 37, "legacy_composites": 54, "chronological_status": 40, "onset_primary_mechanism": 32})
    add_color_scale(event_blocks, ["calendar_mean_ev", "calendar_mean_signed_surprise", "legacy_recognition_rate"], blocks)

    daily = workbook.create_sheet("Daily Calendar")
    write_table(daily, calendar, "DailyCalendarTable", {"archetype_policy_key": 37, "legacy_composites": 54, "recognition_sources": 30, "uncaptured_reason": 40, "onset_primary_mechanism": 32})
    add_color_scale(daily, ["mean_ev_after_1pct", "clean_exec_precision", "daily_ev_neutral_z", "best_composite_score"], calendar)

    focus_ws = workbook.create_sheet("Focus Episodes")
    write_table(focus_ws, focus, "FocusEpisodesTable", {"archetype_policy_key": 37, "recovery_assessment": 48, "research_status": 32, "onset_primary_mechanism": 32, "alerting_families": 34})
    add_color_scale(focus_ws, ["calendar_mean_ev", "sequence_top05_lift", "research_top05_lift", "max_valid_risk"], focus)

    methods_ws = workbook.create_sheet("Method Evidence")
    write_table(methods_ws, methods, "MethodEvidenceTable", {"evaluation_contract": 50, "archetype": 37, "mechanism": 32, "notes": 62})
    add_color_scale(methods_ws, ["lift", "fpr", "recall", "precision"], methods)

    contracts_ws = workbook.create_sheet("Contracts")
    write_contracts(contracts_ws)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
